"""Extract help desk scheduling inputs from Postgres into an Excel workbook.

This script materialises the pieces needed by ``scheduler_lp`` so that the
notebook workflow can run on realistic data without talking to the Flask app.

Two export modes are supported:

* ``pre-schedule`` (default): pull assistants, availability, and courses, then
    synthesise a shift grid for the requested date range. This is useful when you
    want to test a solver before any schedule has been generated in the app.
* ``schedule``: read the latest (or specified) help desk schedule along with its
    shifts and course demand rows.

Usage (PowerShell):

```
python scripts/export_scheduler_data.py --output data/helpdesk_inputs.xlsx
python scripts/export_scheduler_data.py --mode schedule --schedule-id 7
python scripts/export_scheduler_data.py --start-date 2025-01-20 --end-date 2025-01-24 --day-start 8 --day-end 18
```

Environment variables are honoured in the following order to resolve the
connection string: ``DATABASE_URL``, ``SQLALCHEMY_DATABASE_URI``,
``DATABASE_URI_POSTGRES_LOCAL``. You can override them on the command line with
``--database-url``.

The generated Excel workbook contains the following sheets:
 - ``assistants``: Active help desk assistants with minimum hour commitments.
 - ``assistant_courses``: Mapping of assistants to course codes they support.
 - ``assistant_availability``: Weekly availability windows (0 = Monday).
 - ``shifts``: Help desk shifts (synthesised or persisted) with weekday and duration metadata.
 - ``shift_course_demands``: Course coverage requirements per shift.
 - ``metadata``: High level context for the export (mode, date range, schedule id, run time).

Dependencies: SQLAlchemy, psycopg2-binary, and openpyxl must be installed. The
project already includes the first two via Flask-SQLAlchemy; ``openpyxl`` is
added to ``requirements.txt`` by this change.
"""
from __future__ import annotations

import argparse
import os
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Dict, Iterable, Mapping, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Connection
from sqlalchemy.exc import SQLAlchemyError


@dataclass(frozen=True)
class AssistantRow:
    assistant_id: str
    full_name: Optional[str]
    rate: float
    hours_minimum: int
    active: bool


@dataclass(frozen=True)
class AvailabilityRow:
    assistant_id: str
    day_of_week: int
    start_time: str
    end_time: str


@dataclass(frozen=True)
class CourseCapabilityRow:
    assistant_id: str
    course_code: str


@dataclass(frozen=True)
class ShiftRow:
    shift_id: int
    schedule_id: int
    date: str
    day_of_week: int
    start_time: str
    end_time: str
    duration_hours: float


@dataclass(frozen=True)
class ShiftCourseDemandRow:
    shift_id: int
    course_code: str
    tutors_required: int
    weight: int


@dataclass(frozen=True)
class ScheduleInfo:
    schedule_id: int
    start_date: datetime
    end_date: datetime
    generated_at: Optional[datetime]
    is_published: bool


class ExportError(RuntimeError):
    """Raised when the export fails for a user-facing reason."""


def _parse_date(value: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:  # pragma: no cover - argparse surface
        raise argparse.ArgumentTypeError(f"Invalid date '{value}'. Use YYYY-MM-DD format.") from exc


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--database-url",
        dest="database_url",
        help="SQLAlchemy connection string (postgresql://user:pass@host:port/db)",
    )
    parser.add_argument(
        "--mode",
        choices=("pre-schedule", "schedule"),
        default="pre-schedule",
        help="Export mode: synthesise shifts (pre-schedule) or read persisted schedule.",
    )
    parser.add_argument(
        "--schedule-id",
        dest="schedule_id",
        type=int,
        help="Explicit schedule id to export (must be type='helpdesk').",
    )
    parser.add_argument(
        "--output",
        dest="output_path",
        default="helpdesk_scheduler_inputs.xlsx",
        help="Destination .xlsx path (directories will be created as needed).",
    )
    parser.add_argument(
        "--start-date",
        type=_parse_date,
        help="Inclusive start date (YYYY-MM-DD) when synthesising shifts.",
    )
    parser.add_argument(
        "--end-date",
        type=_parse_date,
        help="Inclusive end date (YYYY-MM-DD) when synthesising shifts.",
    )
    parser.add_argument(
        "--day-start",
        type=int,
        default=9,
        help="Hour of day (24h) for the first shift when synthesising shifts.",
    )
    parser.add_argument(
        "--day-end",
        type=int,
        default=17,
        help="Hour of day (24h) for the end of the final shift (exclusive).",
    )
    parser.add_argument(
        "--tutors-required",
        type=int,
        default=2,
        help="Default tutors required per course per shift when synthesising.",
    )
    parser.add_argument(
        "--include-weekends",
        action="store_true",
        help="Include Saturday/Sunday when synthesising shifts.",
    )

    args = parser.parse_args()

    database_url = resolve_database_url(args.database_url)
    if not database_url:
        raise ExportError(
            "No database URL detected. Set DATABASE_URL or pass --database-url."
        )

    engine = create_engine(database_url)

    try:
        with engine.connect() as conn:
            assistants = list(fetch_assistants(conn))
            availabilities = list(fetch_availability(conn))
            capabilities = list(fetch_course_capabilities(conn))
            metadata_extras: Dict[str, str] = {"mode": args.mode}

            if args.mode == "schedule":
                schedule = resolve_schedule(conn, args.schedule_id)
                shifts = list(fetch_shifts(conn, schedule.schedule_id))
                demands = list(fetch_shift_course_demands(conn, schedule.schedule_id))
            else:
                courses = list(fetch_courses(conn))
                if not courses:
                    raise ExportError(
                        "No courses found. Seed the course catalog before exporting pre-schedule data."
                    )
                start_date, end_date = determine_date_range(args.start_date, args.end_date)
                shifts, demands = build_synthetic_shifts(
                    courses,
                    start_date=start_date,
                    end_date=end_date,
                    day_start=args.day_start,
                    day_end=args.day_end,
                    tutors_required=args.tutors_required,
                    include_weekends=args.include_weekends,
                )
                schedule = ScheduleInfo(
                    schedule_id=-1,
                    start_date=datetime.combine(start_date, time.min),
                    end_date=datetime.combine(end_date, time.min),
                    generated_at=None,
                    is_published=False,
                )
                metadata_extras.update(
                    {
                        "synthetic_start_date": start_date.isoformat(),
                        "synthetic_end_date": end_date.isoformat(),
                        "synthetic_day_start": str(args.day_start),
                        "synthetic_day_end": str(args.day_end),
                        "synthetic_include_weekends": str(args.include_weekends),
                        "synthetic_tutors_required": str(args.tutors_required),
                    }
                )
    except SQLAlchemyError as exc:  # pragma: no cover - runtime failure path
        raise ExportError(f"Database query failed: {exc}") from exc

    output_path = Path(args.output_path).expanduser().resolve()
    if not output_path.parent.exists():
        output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    populate_assistants_sheet(workbook, assistants)
    populate_courses_sheet(workbook, capabilities)
    populate_availability_sheet(workbook, availabilities)
    populate_shifts_sheet(workbook, shifts)
    populate_shift_demands_sheet(workbook, demands)
    populate_metadata_sheet(workbook, schedule, database_url, metadata_extras)

    workbook.save(output_path)
    print(f"Exported scheduler inputs to {output_path}")


def resolve_database_url(override: Optional[str]) -> Optional[str]:
    if override:
        return override
    for key in ("DATABASE_URL", "SQLALCHEMY_DATABASE_URI", "DATABASE_URI_POSTGRES_LOCAL"):
        value = os.getenv(key)
        if value:
            return value
    return None


def resolve_schedule(conn: Connection, schedule_id: Optional[int]) -> ScheduleInfo:
    if schedule_id is not None:
        row = conn.execute(
            text(
                """
                SELECT id, start_date, end_date, generated_at, is_published
                FROM schedule
                WHERE id = :schedule_id AND type = 'helpdesk'
                """
            ),
            {"schedule_id": schedule_id},
        ).fetchone()
        if row is None:
            raise ExportError(
                f"Schedule {schedule_id} was not found or is not a help desk schedule."
            )
    else:
        row = conn.execute(
            text(
                """
                SELECT id, start_date, end_date, generated_at, is_published
                FROM schedule
                WHERE type = 'helpdesk'
                ORDER BY generated_at DESC NULLS LAST, id DESC
                LIMIT 1
                """
            )
        ).fetchone()
        if row is None:
            raise ExportError("No help desk schedule exists yet. Use --schedule-id once created.")

    mapping = row._mapping
    return ScheduleInfo(
        schedule_id=mapping["id"],
        start_date=mapping["start_date"],
        end_date=mapping["end_date"],
        generated_at=mapping.get("generated_at"),
        is_published=bool(mapping["is_published"]),
    )


def fetch_assistants(conn: Connection) -> Iterable[AssistantRow]:
    result = conn.execute(
        text(
            """
            SELECT h.username AS assistant_id,
                   h.rate,
                   h.hours_minimum,
                   h.active,
                   s.name AS full_name
            FROM help_desk_assistant h
            JOIN student s ON s.username = h.username
            ORDER BY h.username
            """
        )
    )
    for row in result:
        mapping = row._mapping
        yield AssistantRow(
            assistant_id=mapping["assistant_id"],
            full_name=mapping.get("full_name"),
            rate=float(mapping["rate"]),
            hours_minimum=int(mapping["hours_minimum"]),
            active=bool(mapping["active"]),
        )


def fetch_availability(conn: Connection) -> Iterable[AvailabilityRow]:
    result = conn.execute(
        text(
            """
            SELECT a.username AS assistant_id,
                   a.day_of_week,
                   a.start_time,
                   a.end_time
            FROM availability a
            JOIN help_desk_assistant h ON h.username = a.username
            ORDER BY a.username, a.day_of_week, a.start_time
            """
        )
    )
    for row in result:
        mapping = row._mapping
        yield AvailabilityRow(
            assistant_id=mapping["assistant_id"],
            day_of_week=int(mapping["day_of_week"]),
            start_time=_format_time(mapping["start_time"]),
            end_time=_format_time(mapping["end_time"]),
        )


def fetch_course_capabilities(conn: Connection) -> Iterable[CourseCapabilityRow]:
    result = conn.execute(
        text(
            """
            SELECT cc.assistant_username AS assistant_id,
                   cc.course_code
            FROM course_capability cc
            JOIN help_desk_assistant h ON h.username = cc.assistant_username
            ORDER BY cc.assistant_username, cc.course_code
            """
        )
    )
    for row in result:
        mapping = row._mapping
        yield CourseCapabilityRow(
            assistant_id=mapping["assistant_id"],
            course_code=mapping["course_code"],
        )


def fetch_shifts(conn: Connection, schedule_id: int) -> Iterable[ShiftRow]:
    result = conn.execute(
        text(
            """
            SELECT s.id AS shift_id,
                   s.schedule_id,
                   s.date,
                   s.start_time,
                   s.end_time
            FROM shift s
            WHERE s.schedule_id = :schedule_id
            ORDER BY s.start_time
            """
        ),
        {"schedule_id": schedule_id},
    )
    for row in result:
        mapping = row._mapping
        date_val: datetime = mapping["date"]
        start_dt: datetime = mapping["start_time"]
        end_dt: datetime = mapping["end_time"]
        yield ShiftRow(
            shift_id=int(mapping["shift_id"]),
            schedule_id=int(mapping["schedule_id"]),
            date=date_val.date().isoformat(),
            day_of_week=date_val.weekday(),
            start_time=_format_time(start_dt.time()),
            end_time=_format_time(end_dt.time()),
            duration_hours=round((end_dt - start_dt).total_seconds() / 3600.0, 2),
        )


def fetch_shift_course_demands(conn: Connection, schedule_id: int) -> Iterable[ShiftCourseDemandRow]:
    result = conn.execute(
        text(
            """
            SELECT scd.shift_id,
                   scd.course_code,
                   scd.tutors_required,
                   COALESCE(scd.weight, scd.tutors_required) AS weight
            FROM shift_course_demand scd
            JOIN shift s ON s.id = scd.shift_id
            WHERE s.schedule_id = :schedule_id
            ORDER BY scd.shift_id, scd.course_code
            """
        ),
        {"schedule_id": schedule_id},
    )
    for row in result:
        mapping = row._mapping
        yield ShiftCourseDemandRow(
            shift_id=int(mapping["shift_id"]),
            course_code=mapping["course_code"],
            tutors_required=int(mapping["tutors_required"]),
            weight=int(mapping["weight"]),
        )


def populate_assistants_sheet(workbook: Workbook, rows: Iterable[AssistantRow]) -> None:
    sheet = workbook.active
    sheet.title = "assistants"
    headers = [
        "assistant_id",
        "full_name",
        "rate",
        "hours_minimum",
        "active",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([
            row.assistant_id,
            row.full_name or "",
            row.rate,
            row.hours_minimum,
            int(row.active),
        ])
    autosize_columns(sheet, len(headers))


def populate_courses_sheet(workbook: Workbook, rows: Iterable[CourseCapabilityRow]) -> None:
    sheet = workbook.create_sheet("assistant_courses")
    headers = ["assistant_id", "course_code"]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.assistant_id, row.course_code])
    autosize_columns(sheet, len(headers))


def populate_availability_sheet(workbook: Workbook, rows: Iterable[AvailabilityRow]) -> None:
    sheet = workbook.create_sheet("assistant_availability")
    headers = ["assistant_id", "day_of_week", "start_time", "end_time"]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.assistant_id, row.day_of_week, row.start_time, row.end_time])
    autosize_columns(sheet, len(headers))


def populate_shifts_sheet(workbook: Workbook, rows: Iterable[ShiftRow]) -> None:
    sheet = workbook.create_sheet("shifts")
    headers = [
        "shift_id",
        "schedule_id",
        "date",
        "day_of_week",
        "start_time",
        "end_time",
        "duration_hours",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                row.shift_id,
                row.schedule_id,
                row.date,
                row.day_of_week,
                row.start_time,
                row.end_time,
                row.duration_hours,
            ]
        )
    autosize_columns(sheet, len(headers))


def populate_shift_demands_sheet(workbook: Workbook, rows: Iterable[ShiftCourseDemandRow]) -> None:
    sheet = workbook.create_sheet("shift_course_demands")
    headers = ["shift_id", "course_code", "tutors_required", "weight"]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.shift_id, row.course_code, row.tutors_required, row.weight])
    autosize_columns(sheet, len(headers))


def fetch_courses(conn: Connection) -> Sequence[str]:
    result = conn.execute(
        text(
            """
            SELECT code
            FROM course
            ORDER BY code
            """
        )
    )
    return [row._mapping["code"] for row in result]


def determine_date_range(start: Optional[date], end: Optional[date]) -> Tuple[date, date]:
    if start and end and end < start:
        raise ExportError("End date cannot be earlier than start date.")

    if start is None:
        today = date.today()
        days_until_monday = (7 - today.weekday()) % 7
        start = today + timedelta(days=days_until_monday)
    if end is None:
        end = start + timedelta(days=4)

    return start, end


def build_synthetic_shifts(
    courses: Sequence[str],
    *,
    start_date: date,
    end_date: date,
    day_start: int,
    day_end: int,
    tutors_required: int,
    include_weekends: bool,
) -> Tuple[list[ShiftRow], list[ShiftCourseDemandRow]]:
    if day_end <= day_start:
        raise ExportError("day-end must be greater than day-start")
    if tutors_required <= 0:
        raise ExportError("tutors-required must be positive")

    shift_rows: list[ShiftRow] = []
    demand_rows: list[ShiftCourseDemandRow] = []
    shift_counter = 1

    current = start_date
    while current <= end_date:
        weekday = current.weekday()
        if include_weekends or weekday < 5:
            for hour in range(day_start, day_end):
                start_label = f"{hour:02d}:00"
                end_label = f"{(hour + 1):02d}:00"
                shift_rows.append(
                    ShiftRow(
                        shift_id=shift_counter,
                        schedule_id=0,
                        date=current.isoformat(),
                        day_of_week=weekday,
                        start_time=start_label,
                        end_time=end_label,
                        duration_hours=1.0,
                    )
                )
                for course_code in courses:
                    demand_rows.append(
                        ShiftCourseDemandRow(
                            shift_id=shift_counter,
                            course_code=course_code,
                            tutors_required=tutors_required,
                            weight=tutors_required,
                        )
                    )
                shift_counter += 1
        current += timedelta(days=1)

    return shift_rows, demand_rows


def populate_metadata_sheet(
    workbook: Workbook,
    schedule: ScheduleInfo,
    database_url: str,
    extras: Optional[Mapping[str, str]] = None,
) -> None:
    sheet = workbook.create_sheet("metadata")
    data: Mapping[str, str] = {
        "schedule_id": str(schedule.schedule_id),
        "schedule_start_date": schedule.start_date.isoformat() if schedule.start_date else "",
        "schedule_end_date": schedule.end_date.isoformat() if schedule.end_date else "",
        "schedule_generated_at": schedule.generated_at.isoformat() if schedule.generated_at else "",
        "schedule_published": str(schedule.is_published),
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "database_url": redact_database_url(database_url),
    }
    if extras:
        data = {**data, **extras}
    for key, value in data.items():
        sheet.append([key, value])
    autosize_columns(sheet, 2)


def autosize_columns(sheet, column_count: int) -> None:
    for idx in range(1, column_count + 1):
        column_letter = get_column_letter(idx)
        max_length = 0
        for cell in sheet[column_letter]:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def redact_database_url(url: str) -> str:
    if "@" not in url:
        return url
    _, suffix = url.split("@", 1)
    return "***@" + suffix


def _format_time(value: time) -> str:
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M")
    raise ExportError(f"Unexpected time object {value!r}")


if __name__ == "__main__":  # pragma: no cover - CLI entry point
    try:
        main()
    except ExportError as exc:
        print(f"Export failed: {exc}")
        raise SystemExit(1)
